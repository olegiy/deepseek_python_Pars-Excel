# excel_utils.py

from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

def unmerge_cells_without_filling(ws):
    """Разъединяет объединённые ячейки без заполнения."""
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    logger.debug(f"🔧 Разъединены объединённые ячейки на листе '{ws.title}'.")

def merge_first_row(ws):
    """Объединяет первую строку, если есть данные."""
    if ws.max_row == 0 or ws.max_column == 0:
        logger.debug(f"🚫 Нет данных для объединения первой строки на листе '{ws.title}'.")
        return
    last_col = max((cell.column for row in ws.iter_rows() for cell in row if cell.value), default=0)
    if last_col > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        logger.debug(f"🔗 Объединена первая строка на листе '{ws.title}'.")
        if ws.cell(row=1, column=1).value:
            ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=1, column=1).font = Font(bold=True)

def auto_adjust_column_width(ws):
    """Автоподбор ширины столбцов."""
    for column in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    logger.debug(f"📏 Автоподбор ширины столбцов выполнен для листа '{ws.title}'.")