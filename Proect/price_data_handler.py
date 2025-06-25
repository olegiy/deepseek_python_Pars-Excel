# price_data_handler.py

from tkinter import Tk, filedialog
import openpyxl
from openpyxl.styles import numbers
import os
import logging

logger = logging.getLogger(__name__)
PRICE_SHEET_NAME = "Price Data"

def attach_price_file(wb):
    """Добавляет лист с ценами"""
    logger.info("💲 Выберите файл с ценами (например Price.xlsx)")
    Tk().withdraw()
    price_file = filedialog.askopenfilename(
        title="Выберите файл с ценами",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not price_file:
        logger.warning("🚫 Файл с ценами не выбран. Пропускаем добавление цен.")
        return None

    try:
        price_wb = openpyxl.load_workbook(price_file)
        if not price_wb.sheetnames:
            logger.error("❌ В выбранном файле нет листов с данными.")
            return None

        if PRICE_SHEET_NAME in wb.sheetnames:
            wb.remove(wb[PRICE_SHEET_NAME])

        price_sheet = price_wb[price_wb.sheetnames[0]]
        new_sheet = wb.create_sheet(PRICE_SHEET_NAME)

        for row in price_sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    try:
                        value = float(value.replace(",", ".").replace(" ", ""))
                    except ValueError:
                        pass
                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=value)

                if isinstance(value, (int, float)):
                    new_cell.value = round(value, 2)
                    new_cell.number_format = numbers.FORMAT_NUMBER_00

                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.alignment = cell.alignment.copy()

        for col in price_sheet.column_dimensions:
            new_sheet.column_dimensions[col].width = price_sheet.column_dimensions[col].width

        logger.info(f"✅ Лист с ценами успешно добавлен из файла: {os.path.basename(price_file)}")
        return new_sheet

    except Exception as e:
        logger.error(f"❌ Ошибка при добавлении файла с ценами: {str(e)}")
        return None