# part_info_processor.py

from openpyxl.styles import Font
from pricing import calculate_prices_for_section
import logging

logger = logging.getLogger(__name__)
PART_INFO_SHEET = "Part Info"

def process_part_info_sheet(ws, price_data_ws):
    """Обрабатывает лист Part Info с расчётами"""
    if ws.title != PART_INFO_SHEET:
        return

    logger.info("📄 Обработка листа 'Part Info'")
    # Находим столбец с ценами
    price_col = None
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() in ["price(₽)", "цена"]:
            price_col = cell.column
            break

    if price_col:
        clear_price_column(ws, price_col)

    section_rows = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "section:" in cell.value.lower():
                section_rows.append(cell.row)
                break

    # Обрабатываем секции в обратном порядке
    for i in range(len(section_rows)-1, -1, -1):
        row_num = section_rows[i]
        thickness_values = []
        tube_count = None
        start_data_row = row_num + 1
        end_data_row = ws.max_row if i == len(section_rows)-1 else section_rows[i+1]-1

        # Рассчитываем цены для всех ID в секции и получаем общую стоимость
        total_price_section = calculate_prices_for_section(ws, price_data_ws, row_num, end_data_row)

        header_row = None
        id_col = None
        for r in range(start_data_row, end_data_row + 1):
            row_values = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[r]]
            if "id" in row_values and "part name" in row_values and "qty" in row_values:
                header_row = r
                for cell in ws[r]:
                    if str(cell.value).strip().lower() == "id":
                        id_col = cell.column
                        break
                break

        if not header_row or not id_col:
            continue

        col_indices = {
            "ID": id_col,
            "Part Name": None,
            "Qty": None,
            "Part Length(mm)": None,
            "Contour Qty": None,
            "Cut Length(mm)": None,
            "Price(₽)": None
        }

        for cell in ws[header_row]:
            if cell.value in col_indices:
                col_indices[cell.value] = cell.column

        if not col_indices["Qty"] or not col_indices["Part Length(mm)"]:
            continue

        last_data_row = header_row
        for r in range(header_row + 1, end_data_row + 1):
            id_cell = ws.cell(row=r, column=id_col)
            if id_cell.value is None or (isinstance(id_cell.value, str) and not id_cell.value.strip().isdigit()):
                break
            last_data_row = r

        if last_data_row > header_row:
            total_qty = 0
            total_length = 0
            total_contour = 0
            total_cut_length = 0

            for data_row in range(header_row + 1, last_data_row + 1):
                try:
                    qty = float(ws.cell(row=data_row, column=col_indices["Qty"]).value) or 0
                    part_length = float(ws.cell(row=data_row, column=col_indices["Part Length(mm)"]).value) or 0
                    contour_qty = float(ws.cell(row=data_row, column=col_indices["Contour Qty"]).value) or 0 if col_indices["Contour Qty"] else 0
                    cut_length = float(ws.cell(row=data_row, column=col_indices["Cut Length(mm)"]).value) or 0 if col_indices["Cut Length(mm)"] else 0

                    total_qty += qty
                    total_length += qty * part_length
                    total_contour += qty * contour_qty
                    total_cut_length += qty * cut_length
                except (ValueError, TypeError):
                    continue

            result_row = last_data_row + 1
            ws.insert_rows(result_row)

            if col_indices["Qty"]:
                ws.cell(row=result_row, column=col_indices["Qty"], value=f"Total Qty: {total_qty}")

            if col_indices["Part Length(mm)"]:
                ws.cell(row=result_row, column=col_indices["Part Length(mm)"], value=f"Total Length: {total_length}")

            if col_indices["Contour Qty"] and total_contour > 0:
                ws.cell(row=result_row, column=col_indices["Contour Qty"], value=f"Total Contour: {total_contour}")

            if col_indices["Cut Length(mm)"] and total_cut_length > 0:
                ws.cell(row=result_row, column=col_indices["Cut Length(mm)"], value=f"Total Cut Length: {total_cut_length}")

            # Добавляем общую стоимость секции
            if total_price_section is not None and col_indices["Price(₽)"]:
                ws.cell(row=result_row, column=col_indices["Price(₽)"], value=f"Total Price Section: {total_price_section:.2f}").font = Font(bold=True)

        for cell in ws[row_num]:
            if cell.value and isinstance(cell.value, str):
                cell_text = cell.value.lower()
                for keyword in ["толщина стенки", "средняя толщина ноги", "толщина"]:
                    value = extract_thickness_value(cell_text, keyword)
                    if value is not None:
                        thickness_values.append(value)
                if "tube count:" in cell_text:
                    try:
                        tube_count = int(re.search(r'tube count:\s*(\d+)', cell_text).group(1))
                    except (AttributeError, ValueError):
                        pass

        if thickness_values:
            max_value = math.ceil(max(thickness_values))
            ws.cell(row=row_num, column=5, value=f"Толщина стенки: {max_value}").font = Font(bold=True)

        if tube_count is not None and price_data_ws and thickness_values:
            max_value = math.ceil(max(thickness_values))
            price_per_tube = None

            for price_row in price_data_ws.iter_rows(min_row=2):
                thickness_cell = price_row[0]
                price_cell = price_row[1]

                if thickness_cell.value is not None and price_cell.value is not None:
                    try:
                        thickness = float(thickness_cell.value)
                        if abs(thickness - max_value) < 0.01:
                            price_per_tube = float(price_cell.value)
                            break
                    except ValueError:
                        pass

            if price_per_tube is not None:
                logistics_cost = tube_count * price_per_tube
                ws.cell(row=row_num, column=7, value=f"Logistics Cost: {logistics_cost:.2f}").font = Font(bold=True)

    logger.info("✅ Лист 'Part Info' успешно обработан")