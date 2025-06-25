# formatting.py

from openpyxl.styles import Font, PatternFill, Alignment
import logging

logger = logging.getLogger(__name__)
from config import DARK_RED, LIGHT_YELLOW

def apply_section_row_formatting(ws):
    """Форматирует строки с 'Section:'."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "section:" in cell.value.lower():
                ws.row_dimensions[cell.row].height = 35
                for row_cell in ws[cell.row]:
                    row_cell.font = Font(bold=True)
                    row_cell.alignment = Alignment(horizontal='center', vertical='center')
                logger.debug(f"🎨 Строка секции отформатирована: '{cell.value}' на листе '{ws.title}'.")
                break

def apply_styles_to_sheet(ws):
    """Применяет стили ко всем строкам листа."""
    for row in ws.iter_rows():
        is_section = any(cell.value and isinstance(cell.value, str) and "section:" in str(cell.value).lower() for cell in row)
        is_header = any(cell.value in ["ID", "Part Name", "Qty", "Part Length(mm)", "Contour Qty", "Cut Length(mm)", "Price(₽)"] for cell in row)
        is_result = any(cell.value and isinstance(cell.value, str) and ("total" in str(cell.value).lower() or "logistics cost:" in str(cell.value).lower()) for cell in row)
        is_empty = all(cell.value is None for cell in row)

        if is_section:
            ws.row_dimensions[row[0].row].height = 35
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        elif is_header:
            ws.row_dimensions[row[0].row].height = 25
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        elif is_result:
            ws.row_dimensions[row[0].row].height = 35
            for cell in row:
                cell.font = Font(bold=True, color=DARK_RED)
                cell.fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
        elif is_empty:
            ws.row_dimensions[row[0].row].height = 35
        else:
            ws.row_dimensions[row[0].row].height = 15
            for cell in row:
                cell.font = Font(bold=False)
    logger.debug(f"🎨 Применены стили к листу '{ws.title}'.")