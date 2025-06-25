# pricing.py

import re
import math
from openpyxl.styles import Font, Alignment, PatternFill
from config import DARK_RED, LIGHT_YELLOW
import logging

logger = logging.getLogger(__name__)

def extract_thickness_value(text, keyword):
    """Извлекает значение толщины из текста по ключевому слову"""
    try:
        pattern = re.compile(rf"{re.escape(keyword)}[^\d]*([\d,\.]+)")
        matches = pattern.findall(str(text).lower())
        if matches:
            value = float(matches[-1].replace(',', '.'))
            logger.debug(f"📏 Извлечено значение толщины: {value}")
            return value
    except Exception as e:
        logger.warning(f"⚠️ Не удалось извлечь толщину из текста: {text}, ошибка: {e}")
    return None


def find_closest_price_data(price_ws, thickness):
    """Находит ближайшие данные цены для заданной толщины"""
    closest_data = None
    min_diff = float('inf')
    for row in price_ws.iter_rows(min_row=2):
        try:
            current_thickness = float(row[0].value)
            diff = abs(current_thickness - thickness)
            if diff < min_diff:
                min_diff = diff
                closest_data = {
                    'C': float(row[2].value) if row[2].value else 0,  # Цена за контур
                    'D': float(row[3].value) if row[3].value else 0   # Цена за метр резки
                }
        except (ValueError, TypeError):
            continue
    logger.info(f"🔍 Найдены ближайшие данные цены для толщины: {thickness}")
    return closest_data


def calculate_prices_for_section(ws, price_ws, section_row, next_section_row):
    """Рассчитывает цены для всех ID в секции"""
    logger.info(f"💰 Рассчёт цен для секции начиная со строки {section_row}")
    # Находим толщину стенки и logistics cost для секции
    thickness = None
    logistics_cost = 0

    # Ищем logistics cost в строке секции
    for cell in ws[section_row]:
        if cell.value and isinstance(cell.value, str) and "Logistics Cost:" in cell.value:
            try:
                logistics_cost = float(re.search(r'Logistics Cost:\s*([\d\.]+)', cell.value).group(1))
                logger.debug(f"🚚 Logistics Cost найден: {logistics_cost}")
            except (AttributeError, ValueError) as e:
                logger.error(f"❌ Ошибка при извлечении Logistics Cost: {e}")
                logistics_cost = 0

    # Ищем толщину стенки
    for cell in ws[section_row]:
        if cell.value and isinstance(cell.value, str):
            for keyword in ["толщина стенки", "средняя толщина ноги", "толщина"]:
                value = extract_thickness_value(cell.value, keyword)
                if value is not None:
                    thickness = value
                    break
        if thickness is not None:
            break

    if thickness is None or price_ws is None:
        logger.warning(f"⚠️ Не найдена толщина или Price Data для секции в строке {section_row}")
        return None

    # Находим данные из Price Data
    price_data = find_closest_price_data(price_ws, thickness)
    if not price_data:
        logger.warning(f"⚠️ Не найдены данные цены для толщины {thickness}")
        return None

    # Находим заголовки столбцов
    headers = {}
    for cell in ws[section_row + 1]:
        if cell.value in ["ID", "Part Name", "Qty", "Part Length(mm)", "Contour Qty", "Cut Length(mm)", "Price(₽)"]:
            headers[cell.value] = cell.column

    if not all(key in headers for key in ["ID", "Qty", "Part Length(mm)", "Contour Qty", "Cut Length(mm)", "Price(₽)"]):
        logger.warning("⚠️ Не найдены все необходимые заголовки столбцов")
        return None

    # Рассчитываем Total Length для секции
    total_length = 0
    for row in range(section_row + 2, next_section_row + 1):
        part_length_cell = ws.cell(row=row, column=headers["Part Length(mm)"])
        qty_cell = ws.cell(row=row, column=headers["Qty"])
        try:
            part_length = float(part_length_cell.value) if part_length_cell.value else 0
            qty = float(qty_cell.value) if qty_cell.value else 0
            total_length += part_length * qty
        except (ValueError, TypeError) as e:
            logger.error(f"❌ Ошибка при расчёте total_length: {e}")
            continue

    # Рассчитываем цены для каждого ID и общую стоимость секции
    total_price_section = 0
    for row in range(section_row + 2, next_section_row + 1):
        id_cell = ws.cell(row=row, column=headers["ID"])
        if not id_cell.value:
            continue
        try:
            qty = float(ws.cell(row=row, column=headers["Qty"]).value) if ws.cell(row=row, column=headers["Qty"]).value else 0
            part_length = float(ws.cell(row=row, column=headers["Part Length(mm)"]).value) if ws.cell(row=row, column=headers["Part Length(mm)"]).value else 0
            contour_qty = float(ws.cell(row=row, column=headers["Contour Qty"]).value) if ws.cell(row=row, column=headers["Contour Qty"]).value else 0
            cut_length = float(ws.cell(row=row, column=headers["Cut Length(mm)"]).value) if ws.cell(row=row, column=headers["Cut Length(mm)"]).value else 0

            # Рассчитываем компоненты цены
            contour_cost = contour_qty * price_data['C']
            cut_cost = cut_length / 1000 * price_data['D']
            logistics_part = part_length / total_length * logistics_cost if total_length > 0 else 0

            # Итоговая цена
            price = contour_cost + cut_cost + logistics_part
            ws.cell(row=row, column=headers["Price(₽)"], value=round(price, 2))

            # Добавляем к общей стоимости секции
            total_price_section += price * qty
        except (ValueError, TypeError) as e:
            logger.error(f"❌ Ошибка при расчёте цены для ID {id_cell.value}: {str(e)}")
            ws.cell(row=row, column=headers["Price(₽)"], value="ERROR")

    logger.info(f"✅ Цены для секции рассчитаны. Общая стоимость: {total_price_section:.2f}")
    return round(total_price_section, 2)