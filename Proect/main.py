# main.py

import openpyxl
from file_utils import select_file
from backup_utils import create_backup
from excel_utils import unmerge_cells_without_filling, merge_first_row, auto_adjust_column_width
from data_processing import try_convert
from formatting import apply_styles_to_sheet
import logging

# Настройка логгирования
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("app.log", encoding="utf-8")
    ]
)

logger = logging.getLogger(__name__)

def process_excel(input_path):
    """Основная функция обработки файла."""
    if not input_path:
        logger.warning("⚠️ Файл не выбран.")
        return
    backup_path = create_backup(input_path)
    try:
        wb = openpyxl.load_workbook(input_path)
        logger.info(f"📘 Файл загружен: {input_path}")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            unmerge_cells_without_filling(ws)
            merge_first_row(ws)
            for row in ws.iter_rows():
                for cell in row:
                    converted_value, number_format = try_convert(cell.value)
                    if converted_value is not None:
                        cell.value = converted_value
                        if number_format:
                            cell.number_format = number_format
            apply_styles_to_sheet(ws)
            auto_adjust_column_width(ws)
        wb.save(input_path)
        logger.info(f"💾 Изменения сохранены в исходный файл: {input_path}")
        if backup_path:
            logger.info(f"🔁 Резервная копия создана: {backup_path}")
    except Exception as e:
        logger.error(f"❌ Произошла ошибка при обработке файла: {str(e)}")

if __name__ == "__main__":
    print("📂 Выберите Excel-файл для обработки (.xlsx)")
    selected_file = select_file()
    process_excel(selected_file)