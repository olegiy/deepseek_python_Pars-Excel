import openpyxl
from tkinter import Tk, filedialog
from openpyxl.utils import get_column_letter
import warnings
import os
import math
from openpyxl.styles import Font, PatternFill, Alignment, numbers
import re
import shutil
from datetime import datetime

# Игнорируем предупреждения
warnings.simplefilter("ignore")

# Цвета для стилей
DARK_RED = "8B0000"  # Темно-красный
LIGHT_YELLOW = "FFFF99"  # Светло-желтый

def create_backup(file_path):
    """Создаёт резервную копию файла"""
    backup_dir = os.path.join(os.path.dirname(file_path), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{os.path.splitext(os.path.basename(file_path))[0]}_backup_{timestamp}.xlsx"
    backup_path = os.path.join(backup_dir, backup_name)
    shutil.copy2(file_path, backup_path)
    print(f"✅ Создана резервная копия: {backup_path}")
    return backup_path

def try_convert(value):
    """Преобразует значение в число"""
    if value is None or value == "ERROR:#VALUE!":
        return None, None
    if isinstance(value, (int, float)):
        return (math.ceil(value) if value != int(value) else int(value)), '0'
    if isinstance(value, str):
        value = value.strip()
        if '/' in value and value.replace('/', '', 1).isdigit():
            return int(value.split('/')[0]), '0'
        try:
            num = float(value.replace(',', '.'))
            return (math.ceil(num) if num != int(num) else int(num)), '0'
        except ValueError:
            return value, None
    return value, None

def unmerge_cells_without_filling(ws):
    """Разъединяет объединённые ячейки"""
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))

def apply_section_row_formatting(ws):
    """Форматирует строки с 'Section:'"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "section:" in cell.value.lower():
                ws.row_dimensions[cell.row].height = 35
                for row_cell in ws[cell.row]:
                    row_cell.font = Font(bold=True)
                    row_cell.alignment = Alignment(horizontal='center', vertical='center')
                break

def merge_first_row(ws):
    """Объединяет ячейки первой строки"""
    if ws.max_row == 0 or ws.max_column == 0:
        return
    last_col = max((cell.column for row in ws.iter_rows() for cell in row if cell.value), default=0)
    if last_col > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        if ws.cell(row=1, column=1).value:
            ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=1, column=1).font = Font(bold=True)

def get_section_name(cell_value):
    """Извлекает название секции"""
    if not isinstance(cell_value, str):
        return None
    match = re.search(r'section:\s*(.+?)(?:\s*толщина|thickness|$)', cell_value.lower())
    return match.group(1).strip() if match else None

def copy_tube_counts_to_part_info(wb):
    """Копирует Tube Count с других листов"""
    if "Part Info" not in wb.sheetnames:
        return
        
    section_tube_counts = {}
    
    if "Nesting  Summary" in wb.sheetnames:
        nesting_ws = wb["Nesting  Summary"]
        for row in nesting_ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "section:" in cell.value.lower():
                    section_name = get_section_name(cell.value)
                    if not section_name:
                        continue
                    for offset in range(1, 5):
                        next_cell = nesting_ws.cell(row=cell.row, column=cell.column + offset)
                        if next_cell.value and isinstance(next_cell.value, str) and "tube count:" in next_cell.value.lower():
                            try:
                                tube_count = int(re.search(r'tube count:\s*(\d+)', next_cell.value.lower()).group(1))
                                section_tube_counts[section_name] = tube_count
                            except (AttributeError, ValueError):
                                pass
    
    if not section_tube_counts and "Tube Info" in wb.sheetnames:
        tube_ws = wb["Tube Info"]
        current_section = None
        for row in tube_ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "section:" in cell.value.lower():
                    current_section = get_section_name(cell.value)
                if current_section and cell.column == 1 and isinstance(cell.value, (int, float)):
                    section_tube_counts[current_section] = int(cell.value)
    
    part_info_ws = wb["Part Info"]
    for row in part_info_ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "section:" in cell.value.lower():
                section_name = get_section_name(cell.value)
                if section_name in section_tube_counts:
                    part_info_ws.cell(
                        row=cell.row, 
                        column=6, 
                        value=f"Tube Count: {section_tube_counts[section_name]}"
                    ).font = Font(bold=True)
                    part_info_ws.cell(row=cell.row, column=6).alignment = Alignment(horizontal='center', vertical='center')

def extract_thickness_value(text, keyword):
    """Извлекает значение толщины из текста"""
    try:
        pattern = re.compile(rf"{re.escape(keyword)}[^\d]*([\d,\.]+)")
        matches = pattern.findall(text.lower())
        if matches:
            return float(matches[-1].replace(',', '.'))
    except:
        return None

def clear_price_column(ws, price_col):
    """Очищает столбец с ценами во всех строках"""
    for row in ws.iter_rows():
        if row[price_col-1].value and isinstance(row[price_col-1].value, (int, float, str)):
            if str(row[price_col-1].value).strip().lower() not in ["price(₽)", "цена"]:
                row[price_col-1].value = None

def find_closest_price_data(price_ws, thickness):
    """Находит ближайшие данные цены для заданной толщины"""
    closest_data = None
    min_diff = float('inf')
    
    for row in price_ws.iter_rows(min_row=2):
        try:
            current_thickness = float(row[0].value)
            diff = abs(current_thickness - thickness)
            
            if diff < min_diff:
                min_diff = diff
                closest_data = {
                    'C': float(row[2].value) if row[2].value else 0,  # Цена за контур
                    'D': float(row[3].value) if row[3].value else 0   # Цена за метр резки
                }
        except (ValueError, TypeError):
            continue
    
    return closest_data

def calculate_prices_for_section(ws, price_ws, section_row, next_section_row):
    """Рассчитывает цены для всех ID в секции"""
    # Находим толщину стенки и logistics cost для секции
    thickness = None
    logistics_cost = 0
    
    # Ищем logistics cost в строке секции
    for cell in ws[section_row]:
        if cell.value and isinstance(cell.value, str) and "Logistics Cost:" in cell.value:
            try:
                logistics_cost = float(re.search(r'Logistics Cost:\s*([\d\.]+)', cell.value).group(1))
                print(f"Для секции в строке {section_row} Logistics Cost = {logistics_cost}")
            except (AttributeError, ValueError) as e:
                print(f"Ошибка при извлечении Logistics Cost: {e}")
                logistics_cost = 0
    
    # Ищем толщину стенки
    for cell in ws[section_row]:
        if cell.value and isinstance(cell.value, str):
            for keyword in ["толщина стенки", "средняя толщина ноги", "толщина"]:
                value = extract_thickness_value(cell.value, keyword)
                if value is not None:
                    thickness = value
                    break
        if thickness is not None:
            break
    
    if thickness is None or price_ws is None:
        print(f"Не найдена толщина или Price Data для секции в строке {section_row}")
        return None
    
    # Находим данные из Price Data
    price_data = find_closest_price_data(price_ws, thickness)
    if not price_data:
        print(f"Не найдены данные цены для толщины {thickness}")
        return None
    
    # Находим заголовки столбцов
    headers = {}
    for cell in ws[section_row + 1]:
        if cell.value in ["ID", "Part Name", "Qty", "Part Length(mm)", "Contour Qty", "Cut Length(mm)", "Price(₽)"]:
            headers[cell.value] = cell.column
    
    if not all(key in headers for key in ["ID", "Qty", "Part Length(mm)", "Contour Qty", "Cut Length(mm)", "Price(₽)"]):
        print("Не найдены все необходимые заголовки столбцов")
        return None
    
    # Рассчитываем Total Length для секции
    total_length = 0
    for row in range(section_row + 2, next_section_row + 1):
        part_length_cell = ws.cell(row=row, column=headers["Part Length(mm)"])
        qty_cell = ws.cell(row=row, column=headers["Qty"])
        
        try:
            part_length = float(part_length_cell.value) if part_length_cell.value else 0
            qty = float(qty_cell.value) if qty_cell.value else 0
            total_length += part_length * qty
        except (ValueError, TypeError) as e:
            print(f"Ошибка при расчете total_length: {e}")
            continue
    
    # Рассчитываем цены для каждого ID и общую стоимость секции
    total_price_section = 0
    for row in range(section_row + 2, next_section_row + 1):
        id_cell = ws.cell(row=row, column=headers["ID"])
        if not id_cell.value:
            continue
        
        try:
            qty = float(ws.cell(row=row, column=headers["Qty"]).value) if ws.cell(row=row, column=headers["Qty"]).value else 0
            part_length = float(ws.cell(row=row, column=headers["Part Length(mm)"]).value) if ws.cell(row=row, column=headers["Part Length(mm)"]).value else 0
            contour_qty = float(ws.cell(row=row, column=headers["Contour Qty"]).value) if ws.cell(row=row, column=headers["Contour Qty"]).value else 0
            cut_length = float(ws.cell(row=row, column=headers["Cut Length(mm)"]).value) if ws.cell(row=row, column=headers["Cut Length(mm)"]).value else 0
            
            # Рассчитываем компоненты цены
            contour_cost = contour_qty * price_data['C']
            cut_cost = cut_length / 1000 * price_data['D']
            logistics_part = part_length / total_length * logistics_cost if total_length > 0 else 0
            
            # Выводим в консоль значения для отладки
            print(f"\nРасчет для ID: {id_cell.value}")
            print(f"Длина детали: {part_length} мм")
            print(f"Общая длина секции: {total_length} мм")
            print(f"Logistics Cost секции: {logistics_cost}")
            print(f"Логистическая часть: {part_length}/{total_length}*{logistics_cost} = {logistics_part:.2f}")
            
            # Итоговая цена
            price = contour_cost + cut_cost + logistics_part
            print(f"Итоговая цена: {contour_cost:.2f} (контуры) + {cut_cost:.2f} (резка) + {logistics_part:.2f} (логистика) = {price:.2f}")
                
            # Записываем цену
            ws.cell(row=row, column=headers["Price(₽)"], value=round(price, 2))
            
            # Добавляем к общей стоимости секции
            total_price_section += price * qty
            
        except (ValueError, TypeError) as e:
            print(f"Ошибка при расчете цены для ID {id_cell.value}: {str(e)}")
            ws.cell(row=row, column=headers["Price(₽)"], value="ERROR")
    
    return round(total_price_section, 2)

def apply_styles_to_sheet(ws):
    """Применяет стили форматирования к листу"""
    for row in ws.iter_rows():
        is_section = any(cell.value and isinstance(cell.value, str) and "section:" in str(cell.value).lower() for cell in row)
        is_header = any(cell.value in ["ID", "Part Name", "Qty", "Part Length(mm)", "Contour Qty", "Cut Length(mm)", "Price(₽)"] for cell in row)
        is_result = any(cell.value and isinstance(cell.value, str) and ("total" in str(cell.value).lower() or "logistics cost:" in str(cell.value).lower()) for cell in row)
        is_empty = all(cell.value is None for cell in row)
        
        if is_section:
            ws.row_dimensions[row[0].row].height = 35
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        elif is_header:
            ws.row_dimensions[row[0].row].height = 25
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        elif is_result:
            ws.row_dimensions[row[0].row].height = 35
            for cell in row:
                cell.font = Font(bold=True, color=DARK_RED)
                cell.fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
        elif is_empty:
            ws.row_dimensions[row[0].row].height = 35
        else:
            ws.row_dimensions[row[0].row].height = 15
            for cell in row:
                cell.font = Font(bold=False)

def process_part_info_sheet(ws, price_data_ws):
    """Обрабатывает лист Part Info с расчетами"""
    if ws.title != "Part Info":
        return
    
    # Находим столбец с ценами
    price_col = None
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() in ["price(₽)", "цена"]:
            price_col = cell.column
            break
    
    if price_col:
        # Очищаем ВСЕ значения в столбце Price (кроме заголовков)
        clear_price_column(ws, price_col)
    
    section_rows = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "section:" in cell.value.lower():
                section_rows.append(cell.row)
                break
    
    # Обрабатываем секции в обратном порядке
    for i in range(len(section_rows)-1, -1, -1):
        row_num = section_rows[i]
        thickness_values = []
        tube_count = None
        
        start_data_row = row_num + 1
        end_data_row = ws.max_row if i == len(section_rows)-1 else section_rows[i+1]-1
        
        # Рассчитываем цены для всех ID в секции и получаем общую стоимость
        total_price_section = calculate_prices_for_section(ws, price_data_ws, row_num, end_data_row)
        
        header_row = None
        id_col = None
        for r in range(start_data_row, end_data_row + 1):
            row_values = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[r]]
            if "id" in row_values and "part name" in row_values and "qty" in row_values:
                header_row = r
                for cell in ws[r]:
                    if str(cell.value).strip().lower() == "id":
                        id_col = cell.column
                        break
                break
        
        if not header_row or not id_col:
            continue
            
        col_indices = {
            "ID": id_col,
            "Part Name": None,
            "Qty": None,
            "Part Length(mm)": None,
            "Contour Qty": None,
            "Cut Length(mm)": None,
            "Price(₽)": None
        }
        
        for cell in ws[header_row]:
            if cell.value in col_indices:
                col_indices[cell.value] = cell.column
        
        if not col_indices["Qty"] or not col_indices["Part Length(mm)"]:
            continue
            
        last_data_row = header_row
        for r in range(header_row + 1, end_data_row + 1):
            id_cell = ws.cell(row=r, column=id_col)
            if id_cell.value is None or (isinstance(id_cell.value, str) and not id_cell.value.strip().isdigit()):
                break
            last_data_row = r
        
        if last_data_row > header_row:
            total_qty = 0
            total_length = 0
            total_contour = 0
            total_cut_length = 0
            
            for data_row in range(header_row + 1, last_data_row + 1):
                try:
                    qty = float(ws.cell(row=data_row, column=col_indices["Qty"]).value) or 0
                    part_length = float(ws.cell(row=data_row, column=col_indices["Part Length(mm)"]).value) or 0
                    contour_qty = float(ws.cell(row=data_row, column=col_indices["Contour Qty"]).value) or 0 if col_indices["Contour Qty"] else 0
                    cut_length = float(ws.cell(row=data_row, column=col_indices["Cut Length(mm)"]).value) or 0 if col_indices["Cut Length(mm)"] else 0
                    
                    total_qty += qty
                    total_length += qty * part_length
                    total_contour += qty * contour_qty
                    total_cut_length += qty * cut_length
                except (ValueError, TypeError):
                    continue
            
            result_row = last_data_row + 1
            ws.insert_rows(result_row)
            
            if col_indices["Qty"]:
                ws.cell(row=result_row, column=col_indices["Qty"], value=f"Total Qty: {total_qty}")
            if col_indices["Part Length(mm)"]:
                ws.cell(row=result_row, column=col_indices["Part Length(mm)"], value=f"Total Length: {total_length}")
            if col_indices["Contour Qty"] and total_contour > 0:
                ws.cell(row=result_row, column=col_indices["Contour Qty"], value=f"Total Contour: {total_contour}")
            if col_indices["Cut Length(mm)"] and total_cut_length > 0:
                ws.cell(row=result_row, column=col_indices["Cut Length(mm)"], value=f"Total Cut Length: {total_cut_length}")
            
            # Добавляем общую стоимость секции
            if total_price_section is not None and col_indices["Price(₽)"]:
                ws.cell(row=result_row, column=col_indices["Price(₽)"], value=f"Total Price Section: {total_price_section:.2f}")
        
        for cell in ws[row_num]:
            if cell.value and isinstance(cell.value, str):
                cell_text = cell.value.lower()
                for keyword in ["толщина стенки", "средняя толщина ноги", "толщина"]:
                    value = extract_thickness_value(cell_text, keyword)
                    if value is not None:
                        thickness_values.append(value)
                if "tube count:" in cell_text:
                    try:
                        tube_count = int(re.search(r'tube count:\s*(\d+)', cell_text).group(1))
                    except (AttributeError, ValueError):
                        pass
        
        if thickness_values:
            max_value = math.ceil(max(thickness_values))
            ws.cell(row=row_num, column=5, value=f"Толщина стенки: {max_value}").font = Font(bold=True)
        
        if tube_count is not None and price_data_ws and thickness_values:
            max_value = math.ceil(max(thickness_values))
            price_per_tube = None
            for price_row in price_data_ws.iter_rows(min_row=2):
                thickness_cell = price_row[0]
                price_cell = price_row[1]
                if thickness_cell.value is not None and price_cell.value is not None:
                    try:
                        thickness = float(thickness_cell.value)
                        if abs(thickness - max_value) < 0.01:
                            price_per_tube = float(price_cell.value)
                            break
                    except ValueError:
                        pass
            
            if price_per_tube is not None:
                logistics_cost = tube_count * price_per_tube
                ws.cell(row=row_num, column=7, value=f"Logistics Cost: {logistics_cost:.2f}").font = Font(bold=True)

def attach_price_file(wb):
    """Добавляет лист с ценами"""
    print("\n💲 Выберите файл с ценами (например Price.xlsx)")
    root = Tk()
    root.withdraw()
    price_file = filedialog.askopenfilename(
        title="Выберите файл с ценами",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not price_file:
        print("Файл с ценами не выбран. Пропускаем добавление цен.")
        return None
    
    try:
        price_wb = openpyxl.load_workbook(price_file)
        if not price_wb.sheetnames:
            print("❌ В выбранном файле нет листов с данными.")
            return None
            
        if "Price Data" in wb.sheetnames:
            wb.remove(wb["Price Data"])
        
        price_sheet = price_wb[price_wb.sheetnames[0]]
        new_sheet = wb.create_sheet("Price Data")
        
        for row in price_sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    try:
                        value = float(value.replace(",", ".").replace(" ", ""))
                    except ValueError:
                        pass
                
                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=value)
                
                if isinstance(value, (int, float)):
                    new_cell.value = round(value, 2)
                    new_cell.number_format = numbers.FORMAT_NUMBER_00
                
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.alignment = cell.alignment.copy()
        
        for col in price_sheet.column_dimensions:
            new_sheet.column_dimensions[col].width = price_sheet.column_dimensions[col].width
        
        print(f"✅ Лист с ценами успешно добавлен из файла: {os.path.basename(price_file)}")
        return new_sheet
        
    except Exception as e:
        print(f"❌ Ошибка при добавлении файла с ценами: {str(e)}")
        return None

def select_file():
    """Выбор файла для обработки"""
    Tk().withdraw()
    return filedialog.askopenfilename(
        title="Выберите Excel-файл",
        filetypes=[("Excel files", "*.xlsx")]
    )

def process_excel(input_path):
    """Основная функция обработки файла"""
    if not input_path:
        print("Файл не выбран.")
        return
    
    backup_path = create_backup(input_path)
    try:
        wb = openpyxl.load_workbook(input_path)
        
        copy_tube_counts_to_part_info(wb)
        price_data_ws = attach_price_file(wb)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            unmerge_cells_without_filling(ws)
            merge_first_row(ws)
            
            for row in ws.iter_rows():
                for cell in row:
                    converted_value, number_format = try_convert(cell.value)
                    if converted_value is not None:
                        cell.value = converted_value
                        if number_format:
                            cell.number_format = number_format
            
            if sheet_name == "Part Info":
                process_part_info_sheet(ws, price_data_ws)
            
            apply_styles_to_sheet(ws)
            
            for column in ws.columns:
                max_length = max(
                    (len(str(cell.value)) for cell in column if cell.value),
                    default=0
                )
                adjusted_width = max_length + 2
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        wb.save(input_path)
        print(f"\n💾 Изменения сохранены в исходный файл: {input_path}")
        print(f"🔁 Резервная копия создана: {backup_path}")
        
    except Exception as e:
        print(f"❌ Произошла ошибка при обработке файла: {str(e)}")
        print("⚠️ Изменения не сохранены. Используйте резервную копию при необходимости.")

if __name__ == "__main__":
    print("📂 Выберите Excel-файл для обработки (.xlsx)")
    selected_file = select_file()
    process_excel(selected_file)